from openpyxl import load_workbook
import re

wb = load_workbook('서울특별시 종합병원 정보.xlsx')
data = wb.active

print('─'*200)

######################################################
# 병원 정보 가져와서 사업장명 / 도로명주소 뽑아내기

names = []
addr = []
for row in data.iter_rows(min_col=2, max_col=2):
    for cell in row:
        # print(cell.value)
        names.append(cell)
        # rows.append(datas)
for row in data.iter_rows(min_col=4, max_col=4):
    for cell in row:
        addr.append(cell)

hosaddr = []
for i in range(1,len(addr)):
    hosaddr.append([names[i].value , addr[i].value])
    # print('[',names[i].value,'] 도로명주소 >>>>> [',addr[i].value,']')

# print(hosaddr)

# ws = wb.create_sheet('서울 종합병원 주소')

# ws['a1'] = '사업장명'
# ws['b1'] = '도로명주소'

# for i in range(1,len(hosaddr)):
#     ws.append(hosaddr[i])

# wb.save('서울특별시 종합병원 정보.xlsx')

######################################################
# 병원을 지역구별로 나누기

'''
도로명주소에서
서울특별시 @@구 ##로 ~~
@@구만 저장할 수 있게 필터링해서(re.compile)
시트로 저장

그리고 지역구 re.compile로 뽑아낸 정보를
해당 시트에 저장하기

>>로직은 다똑같으니 for문?
[지역구for문 = 중복이 있을 수 있으니 중복안되게 단일(셋에 담아서 for문 돌리기)]:
    해당하는 지역구 이름으로 시트생성
    [병원데이터 for문]:
        [if 지역 컴파일러로 일치하는 데이터면]:
            생성된 시트에 데이터 추가
파일에 저장
'''

# 지역구 컴파일
localptn = re.compile('(=? ).+구 ') 
# print(localptn.search('서울특별시 김땡구 구구링'))

local = set()
for row in data.iter_rows(min_col=4, max_col=4):
    for cell in row:
        v = set(re.findall(' .+구 ', cell.value))
        local.update(v)

# for i in addr:
#     v = set(re.findall(' .+구 ', cell.value))
#     local.update(v)


local = list(local)
# print(local,type(local))

for i in range(1, len(local)):
    print(local[i])
    ws = wb.create_sheet(local[i])
    for i in range(1,len(addr)):        
        if localptn.search(names[i].value):
            hosaddr.append([names[i].value , addr[i].value])
    
# print(hosaddr)
print(local,type(local))




wb.save('서울특별시 종합병원 정보.xlsx')
